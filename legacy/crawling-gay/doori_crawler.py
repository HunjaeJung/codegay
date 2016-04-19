# -*- coding:utf-8 -*-
from bs4 import BeautifulSoup
from urllib.request import urlopen
import sys
import csv
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter


"""

# Title         : Get exhibitors list and detail information
# Author        : Hunjae Jung
# Date          : 2015-10-21
# Description   :

- Job 1
총 1페이지부터 365페이지까지 돌면서, Exhibitor와 Country, 그리고 Job 2에서 사용할 kid를 받아옵니다.
start +20씩* 365번

ex)
http://www.anuga.com/anuga/exhibitor-search/search/index.php?fw_goto=aussteller/blaettern&&start=80&paginatevalues=[]
http://www.anuga.com/anuga/exhibitor-search/search/index.php?fw_goto=aussteller/blaettern&&start=120&paginatevalues=[]
http://www.anuga.com/anuga/exhibitor-search/search/index.php?fw_goto=aussteller/blaettern&&start=7300&paginatevalues=[]

- Job 2
Job 1의 결과물을 base로 Exhibitor의 detail information을 받아옵니다.

ex)
http://www.anuga.com/anuga/exhibitor-search/search/index.php?fw_goto=aussteller/details&&kid=0040194960
http://www.anuga.com/anuga/exhibitor-search/search/index.php?fw_goto=aussteller/details&&kid=0002381228
http://www.anuga.com/anuga/exhibitor-search/search/index.php?fw_goto=aussteller/details&&kid=0000988243

# How to run this program
1. python3 doori_crawler.py job1
It will produce ./exhibitor-list.csv

2. python3 doori_crawler.py job2 {{from}} {{to}}
It will produce ./result-detail-{{from}}-{{to}}.csv and ./result-detail-{{from}}-{{to}}.xlsx

"""


def save_spreadsheet(filename, data):
    wb = Workbook()
    ws = wb.active

    row_index = 1
    for rows in data:
        col_index = 1

        for field in rows:
            col_letter = get_column_letter(col_index)
            ws.cell('{}{}'.format(col_letter, row_index)).value = field
            col_index += 1
        row_index += 1

    wb.save(filename)


def do_crawl_job1():
    base_url = "http://www.anuga.com/anuga/exhibitor-search/search/index.php?fw_goto=aussteller/blaettern&&start={{pageId}}&paginatevalues=[]"
    file_name = "./exhibitor-list.csv"

    with open(file_name, 'w') as f:
        writer = csv.writer(f, csv.excel)

        for i in range(0,7301,20):
            url = base_url.replace("{{pageId}}", str(i))
            print('crawling.. => '+url)

            soup = BeautifulSoup(urlopen(url).read())

            exhibitors_list = soup.select('.ergebnisliste > tr > td')

            for i in range(int(len(exhibitors_list)/5)):
                exhibitor_name = exhibitors_list[i*5+1].get_text(' ').strip()
                exhibitor_href = exhibitors_list[i*5+1].select('a')[0].get('href')
                exhibitor_kid = exhibitor_href.split('kid=')[1].split('&')[0]
                exhibitors_addr = exhibitors_list[i*5+2].get_text().strip()

                row = [exhibitor_name, exhibitor_kid, exhibitors_addr]

                writer.writerow(row)

def do_crawl_job2(start_line, end_line):
    base_url = "http://www.anuga.com/anuga/exhibitor-search/search/index.php?fw_goto=aussteller/details&&kid={{kid}}"
    job1_file_name = "./exhibitor-list.csv"
    file_name = "./result-detail-"+str(start_line)+"-"+str(end_line)+".csv"
    failed_list = "./result-failed-"+str(start_line)+"-"+str(end_line)+".csv"

    data = []
    with open(job1_file_name, 'r') as f:
        reader = csv.reader(f, csv.excel)
        for row in reader:
            data.append(row)

    data = data[start_line:end_line]

    with open(failed_list, 'w') as failed_f:
        fail_writer = csv.writer(failed_f, csv.excel)

        with open(file_name, 'w') as f:
            writer = csv.writer(f, csv.excel)

            idx = 0
            for info in data:
                idx = idx + 1
                print(str(idx) + "/" +  str(len(data)) + " crawling.. from " + str(start_line) + " to " + str(end_line))

                try:
                    name = info[0]
                    country = info[2]
                    kid = info[1]

                    url = base_url.replace("{{kid}}", kid)
                    soup = BeautifulSoup(urlopen(url).read())

                    info_detail = " ".join(soup.select("div.widthhalf > div")[2].get_text().replace('\t','').replace('\n','').strip().split())

                    addr = info_detail.split('Phone')[0].strip()

                    if "Fax" in info_detail:
                        phone = info_detail.split('Phone:')[1].split('Fax')[0].strip()
                    else:
                        phone = info_detail.split('Phone:')[1].split('Email')[0].strip()

                    if "Email" in info_detail:
                        email = info_detail.split('Email:')[1].split('Website')[0].strip()
                    else:
                        email = ""

                    if "Website" in info_detail:
                        website = info_detail.split('Website:')[1].strip().split(' ')[0].strip()
                    else:
                        website = ""

                    product_detail = soup.select("div.widthfull > div > div > div > ul.ultree")
                    product_list = product_detail[0].select('> li')
                    trade_show = product_detail[1].get_text().strip()

                    big_categories = []
                    middle_categories = []
                    small_categories = []

                    # 대분류
                    for product in product_list:
                        big_one_name = product.select('a')[0].get_text().strip()
                        big_categories.append(big_one_name)

                        # 중분류
                        middle_list = product.select('> ul')
                        for middle_one in middle_list:
                            middle_one_name = middle_one.select('a')[0].get_text().strip()
                            middle_categories.append(middle_one_name)

                            # 소분류
                            small_list = middle_one.select('ul')
                            for small_one in small_list:
                                small_one_name = small_one.select('a')[0].get_text().strip()
                                small_categories.append(small_one_name)

                    big_categories_str = ", ".join(big_categories)
                    middle_categories_str = ", ".join(middle_categories)
                    small_categories_str = ", ".join(small_categories)

                    row = [name, country, addr, phone, email, website, big_categories_str, middle_categories_str, small_categories_str, trade_show, url]

                    writer.writerow(row)
                except:
                    row = [str(idx), url]
                    fail_writer.writerow(row)
                    print(str(idx) + "/" +  str(len(data)) + " FAILED: " + url)

    return file_name


def do_crawl_save(job2_file_name):
    data = []
    with open(job2_file_name, 'r') as f:
        reader = csv.reader(f, csv.excel)
        for row in reader:
            data.append(row)

    result_file = job2_file_name.split('.csv')[0] + ".xlsx"
    save_spreadsheet(result_file, data)
    print(result_file + " file is succesfully created.")


if __name__ == "__main__":
    job_type = sys.argv[1]

    if job_type == "job1":
        do_crawl_job1()
    elif job_type == "job2":
        start_line = int(sys.argv[2])
        end_line = int(sys.argv[3])
        file_name = do_crawl_job2(start_line, end_line)
        do_crawl_save(file_name)
    else:
        do_crawl_save(sys.argv[2])