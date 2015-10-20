# -*- coding:utf-8 -*-
from bs4 import BeautifulSoup
from urllib.request import urlopen
import sys
import csv
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter


"""

# Title         : Get company information
# Author        : Hunjae Jung
# Date          : 2015-10-20
# Description   :

- Job 1
총 1페이지부터 365페이지까지 돌면서, Exhibitor와 Country, 그리고 Job 2에서 사용할 kid를 받아옵니다.
start +20씩* 365번

ex)
http://www.anuga.com/anuga/exhibitor-search/search/index.php?fw_goto=aussteller/blaettern&&start=80&paginatevalues=[]
http://www.anuga.com/anuga/exhibitor-search/search/index.php?fw_goto=aussteller/blaettern&&start=120&paginatevalues=[]
http://www.anuga.com/anuga/exhibitor-search/search/index.php?fw_goto=aussteller/blaettern&&start=7300&paginatevalues=[]

- Job 2
Job 1의 결과물을 base로 Exhibitor의 detail information을 받아옵니다.

ex)
http://www.anuga.com/anuga/exhibitor-search/search/index.php?fw_goto=aussteller/details&&kid=0040194960
http://www.anuga.com/anuga/exhibitor-search/search/index.php?fw_goto=aussteller/details&&kid=0002381228
http://www.anuga.com/anuga/exhibitor-search/search/index.php?fw_goto=aussteller/details&&kid=0000988243

# How to run this program
1. python3 doori_crawler.py job1
It will produce ./exhibitor-list.csv

2. python3 doori_crawler.py job2
It will produce ./exhibitor-detail.csv

"""


def do_crawl(job_type):
    if job_type == "job1":
        base_url = "http://www.anuga.com/anuga/exhibitor-search/search/index.php?fw_goto=aussteller/blaettern&&start={{pageId}}&paginatevalues=[]"
        file_name = "./exhibitor-list.csv"

        with open(file_name, 'w') as f:
            writer = csv.writer(f, csv.excel)

            for i in range(0,7301,20):
                url = base_url.replace("{{pageId}}", str(i))
                print('crawling.. => '+url)

                soup = BeautifulSoup(urlopen(url).read())

                exhibitors_list = soup.select('.ergebnisliste > tr > td')

                for i in range(int(len(exhibitors_list)/5)):
                    exhibitor_name = exhibitors_list[i*5+1].get_text(' ').strip()
                    exhibitor_href = exhibitors_list[i*5+1].select('a')[0].get('href')
                    exhibitor_kid = exhibitor_href.split('kid=')[1].split('&')[0]
                    exhibitors_addr = exhibitors_list[i*5+2].get_text().strip()

                    row = [exhibitor_name, exhibitor_kid, exhibitors_addr]

                    writer.writerow(row)

    elif job_type == "job2":
        base_url = "http://www.anuga.com/anuga/exhibitor-search/search/index.php?fw_goto=aussteller/details&&kid={{kid}}"
        job1_file_name = "./exhibitor-list.csv"
        file_name = "./exhibitor-detail.csv"

        data = []
        with open(job1_file_name, 'r') as f:
            reader = csv.reader(f, csv.excel)
            for row in reader:
                data.append(row)

        with open(file_name, 'w') as f:
            writer = csv.writer(f, csv.excel)

            idx = 0
            for info in data:
                idx = idx + 1
                print(str(idx) + "/" +  str(len(data)) + " crawling..")

                if idx==5:
                    return

                name = info[0]
                country = info[2]
                kid = info[1]

                url = base_url.replace("{{kid}}", kid)
                soup = BeautifulSoup(urlopen(url).read())

                info_detail = " ".join(soup.select("div.widthhalf > div")[2].get_text().replace('\t','').replace('\n','').strip().split())
                addr = info_detail.split('Phone')[0].strip()
                phone = info_detail.split('Phone:')[1].split('Fax')[0].strip()
                email = info_detail.split('Email:')[1].split('Website')[0].strip()
                website = info_detail.split('Website:')[1].strip().split(' ')[0].strip()

                product_detail = soup.select("div.widthfull > div > div > div > ul.ultree")
                product_list = product_detail[0].select('> li')
                trade_show = product_detail[1].get_text().strip()

                big_categories = []
                middle_categories = []
                small_categories = []

                # 대분류
                for product in product_list:
                    big_one_name = product.select('a')[0].get_text().strip()
                    big_categories.append(big_one_name)

                    # 중분류
                    middle_list = product.select('> ul')
                    for middle_one in middle_list:
                        middle_one_name = middle_one.select('a')[0].get_text().strip()
                        middle_categories.append(middle_one_name)

                        # 소분류
                        small_list = middle_one.select('ul')
                        for small_one in small_list:
                            small_one_name = small_one.select('a')[0].get_text().strip()
                            small_categories.append(small_one_name)

                big_categories_str = ", ".join(big_categories)
                middle_categories_str = ", ".join(middle_categories)
                small_categories_str = ", ".join(small_categories)

                row = [name, country, addr, phone, email, website, big_categories_str, middle_categories_str, small_categories_str, trade_show]

                writer.writerow(row)

    else:
        job2_file_name = "./exhibitor-detail.csv"

        data = []
        with open(job2_file_name, 'r') as f:
            reader = csv.reader(f, csv.excel)
            for row in reader:
                data.append(row)

        print(data[0])
        save_spreadsheet("exhibitors.xlsx", data)


def save_spreadsheet(filename, data):
    wb = Workbook()
    ws = wb.active

    row_index = 1
    for rows in data:
        col_index = 1

        for field in rows:
            col_letter = get_column_letter(col_index)
            ws.cell('{}{}'.format(col_letter, row_index)).value = field
            col_index += 1
        row_index += 1

    wb.save(filename)

if __name__ == "__main__":
    do_crawl(sys.argv[1])
